import ast

import numpy
import openpyxl
from phoBert import cosin_simil, get_10_max
from read_dtu_example import write_verb_example


def read_excel(pathFile, sheet):
    wb = openpyxl.load_workbook(pathFile)
    sheet = wb[sheet]
    max_row = sheet.max_row

    defin = []
    example = []
    verb = []
    vector = []
    for row in range(2, max_row + 1):
        str_df = str(sheet[f"A{row}"].value).split("\n")[0]
        examp = str(sheet[f"B{row}"].value).split("\n")[0]
        vb = str(sheet[f"C{row}"].value).split("_")[0]
        vt = str(sheet[f"D{row}"].value).split("\n")[0]
        vector.append(vt)
        defin.append(str_df)
        example.append(examp)
        verb.append(vb)

    return defin, example, verb, vector


def write_verbnet_vcl(verb_vcl, vectors_vcl, examples_vcl, definitions_vcl,
                      verb_verbnet, vectors_verbnet, definitions_verbnet):
    sheet_idx = int(len(verb_vcl) / 200) + 1
    wb = openpyxl.load_workbook("result_vcl_verb_net.xlsx")
    #
    for k in range(55, sheet_idx):
        print("*******************************************")
        print(k)
        start = (k - 1) * 200
        end = k * 200 if k * 200 < len(verb_vcl) else len(verb_vcl)
        sheet_name = "Sheet_" + str(k)
        cosins = []

        for i in range(start, end):
            cosin_word = []
            for j in range(len(verb_verbnet)):
                cosin = cosin_simil(ast.literal_eval(vectors_vcl[i]), ast.literal_eval(vectors_verbnet[j]))
                cosin_word.append(str(round(cosin, 4)))
            top_10_idx = get_10_max(cosin_word)

            cosin_10_max = {}
            for k in top_10_idx:
                cosin_10_max[definitions_verbnet[k]] = str(cosin_word[k])

            cosins.append(str(cosin_10_max))
        # print(type(str(cosin_10_max)))
        print(numpy.array(cosins).shape)
        write_verb_example("result_vcl_verb_net.xlsx", sheet_name, wb, verb_vcl[start:end],
                           examples_vcl[start:end], definitions_vcl[start:end],
                           vectors_vcl[start:end], cosins)


# defin_vn, example_vn, verb_vn, vector_vn = read_excel("verb_english_bert.xlsx", "Sheet1")
# defin_vcl, example_vcl, verb_vcl, vector_vcl = read_excel("verb_vcl_bert.xlsx", "Sheet1")
# print(".........................")
# write_verbnet_vcl(verb_vcl, vector_vcl, example_vcl, defin_vcl, verb_vn, vector_vn, defin_vn)
