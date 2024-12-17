import numpy as np
import openpyxl
import spacy
from transformers import BertTokenizer, BertModel
from read_dtu_example import  write_verb_example

tokenizer = BertTokenizer.from_pretrained('bert-base-multilingual-cased')
model = BertModel.from_pretrained("bert-base-multilingual-cased")


def read_excel(pathFile, sheet):
    wb = openpyxl.load_workbook(pathFile)
    sheet = wb[sheet]
    max_row = sheet.max_row

    defin_vietnam = []
    example_vietnam = []
    verb = []
    for row in range(2, max_row + 1):
        str_vn = str(sheet[f"A{row}"].value).split("\n")[0]
        examp = str(sheet[f"B{row}"].value).split("\n")[0]
        vb = str(sheet[f"C{row}"].value).split("_")[0]
        defin_vietnam.append(str_vn)
        example_vietnam.append(examp)
        verb.append(vb)

    return defin_vietnam, example_vietnam, verb


def get_vector(word, text):
    tokens = tokenizer.tokenize(tokenizer.decode(tokenizer.encode(text)))
    # print("=====================================================")
    # print(tokens)
    # print(word)
    start = 0
    end = 0
    encoded_input = tokenizer(text, return_tensors='pt')
    for tk in tokens:
        if word.startswith(tk.replace("##", "")):
            start = tokens.index(tk)
        if word.endswith(tk.replace("##", "")):
            end = tokens.index(tk)

    outputs = model(**encoded_input)
    # print(encoded_input)
    target_word_embedding = outputs.last_hidden_state[:, start, :]
    for i in range(start + 1, end + 1):
        target_word_embedding += outputs.last_hidden_state[:, i, :]
    print(np.array(target_word_embedding.tolist()).shape)
    return target_word_embedding.tolist()[0]


def get_vector_vcl(verbs, texts):
    vts = []
    for i in range(len(verbs)):
        vector = get_vector(verbs[i], texts[i])
        print(len(vector))
        vts.append(str(vector))
    return vts


# defin, examples, verbs = read_excel("verb_vn_phoBert.xlsx", "Sheet1")
# print(len(defin), len(examples), len(verbs))
# vectors = get_vector_vcl(verbs, examples)
# print(len(vectors))
# wb = openpyxl.Workbook()
# write_verb_example("verb_vcl_bert.xlsx", "Sheet1", wb, verbs, examples, defin, vectors, [])
# words, examples = read_excel("VerbNet-VietAnh-2022.xlsx", "LeCuong")
# print(words)
# print("=============================")
# print(examples)
